import os
import time
import threading
import logging
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from django.conf import settings

logger = logging.getLogger(__name__)
excel_lock = threading.Lock()

EXPORT_DIR = os.path.join(settings.BASE_DIR, 'export')
os.makedirs(EXPORT_DIR, exist_ok=True)
FILEPATH = os.path.join(EXPORT_DIR, 'Grocery_Online_Data.xlsx')

SHEETS_CONFIG = {
    'Categories': {
        'headers': ['Category ID', 'Category Name', 'Description'],
        'get_row_data': lambda obj: [obj.id, obj.name, obj.description],
        'id_col_idx': 1,
    },
    'Products': {
        'headers': ['Product ID', 'Product Name', 'Category', 'Description', 'Price', 'Stock', 'Rating', 'Created Date'],
        'get_row_data': lambda obj: [
            obj.id,
            obj.name,
            obj.category.name if obj.category else '',
            obj.description,
            float(obj.price) if obj.price is not None else 0.0,
            obj.stock,
            float(obj.rating) if obj.rating is not None else 5.0,
            obj.created_at.strftime('%Y-%m-%d %H:%M:%S') if obj.created_at else ''
        ],
        'id_col_idx': 1,
        'formats': {5: '"₹"#,##0.00', 7: '0.0'},
    },
    'Users': {
        'headers': ['User ID', 'Username', 'Email', 'Date Joined', 'Active Status'],
        'get_row_data': lambda obj: [
            obj.id,
            obj.username,
            obj.email,
            obj.date_joined.strftime('%Y-%m-%d %H:%M:%S') if obj.date_joined else '',
            'Active' if obj.is_active else 'Inactive'
        ],
        'id_col_idx': 1,
    },
    'Orders': {
        'headers': ['Order ID', 'User', 'Order Date', 'Total Amount', 'Payment Status', 'Order Status'],
        'get_row_data': lambda obj: [
            obj.id,
            obj.user.username if obj.user else '',
            obj.created_at.strftime('%Y-%m-%d %H:%M:%S') if obj.created_at else '',
            float(obj.total_amount) if obj.total_amount is not None else 0.0,
            obj.payment_status,
            obj.get_status_display()
        ],
        'id_col_idx': 1,
        'formats': {4: '"₹"#,##0.00'},
    },
    'Order_Items': {
        'headers': ['Order Item ID', 'Order ID', 'Product', 'Quantity', 'Price', 'Subtotal'],
        'get_row_data': lambda obj: [
            obj.id,
            obj.order.id if obj.order else '',
            obj.product_name,
            obj.quantity,
            float(obj.price) if obj.price is not None else 0.0,
            float(obj.get_subtotal()) if obj.get_subtotal() is not None else 0.0
        ],
        'id_col_idx': 1,
        'formats': {5: '"₹"#,##0.00', 6: '"₹"#,##0.00'},
    },
    'Carts': {
        'headers': ['Cart ID', 'User', 'Product', 'Quantity', 'Updated Date'],
        'get_row_data': lambda obj: [
            obj.id,
            obj.user.username if obj.user else '',
            obj.product.name if obj.product else '',
            obj.quantity,
            obj.added_at.strftime('%Y-%m-%d %H:%M:%S') if obj.added_at else ''
        ],
        'id_col_idx': 1,
    },
    'Wishlists': {
        'headers': ['Wishlist ID', 'User', 'Product', 'Added Date'],
        'get_row_data': lambda obj: [
            obj.id,
            obj.user.username if obj.user else '',
            obj.product.name if obj.product else '',
            obj.added_at.strftime('%Y-%m-%d %H:%M:%S') if obj.added_at else ''
        ],
        'id_col_idx': 1,
    }
}

def format_row_cells(ws, row_idx, config):
    formats = config.get('formats', {})
    for col_idx, fmt in formats.items():
        cell = ws.cell(row=row_idx, column=col_idx)
        cell.number_format = fmt

def adjust_column_widths(ws):
    if ws.max_column > 0:
        ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}1"
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            val = str(cell.value or '')
            if len(val) > max_len:
                max_len = len(val)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

def initialize_excel(wb=None):
    if wb is None:
        wb = openpyxl.Workbook()
    
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])
        
    for sheet_name, config in SHEETS_CONFIG.items():
        if sheet_name not in wb.sheetnames:
            ws = wb.create_sheet(title=sheet_name)
            ws.append(config['headers'])
            
            header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
            header_fill = PatternFill(start_color='10B981', end_color='10B981', fill_type='solid')
            
            for col_idx in range(1, len(config['headers']) + 1):
                cell = ws.cell(row=1, column=col_idx)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            ws.freeze_panes = 'A2'
            
    return wb

def save_workbook_with_retry(wb, filepath, max_retries=3, delay=0.2):
    for attempt in range(max_retries):
        try:
            wb.save(filepath)
            return True
        except PermissionError as e:
            logger.warning(f"Failed to save Excel file on attempt {attempt+1}: {str(e)}")
            if attempt < max_retries - 1:
                time.sleep(delay)
            else:
                logger.error("Excel file is open/locked and cannot be saved.")
                raise e
        except Exception as e:
            logger.error(f"Error saving workbook: {str(e)}")
            raise e
    return False

def sync_model_to_excel(sheet_name, instance, action):
    try:
        config = SHEETS_CONFIG.get(sheet_name)
        if not config:
            return

        with excel_lock:
            wb = None
            if os.path.exists(FILEPATH):
                try:
                    wb = openpyxl.load_workbook(FILEPATH)
                except Exception as e:
                    logger.warning(f"Error loading workbook, creating new one: {str(e)}")
            
            if wb is None:
                wb = openpyxl.Workbook()
            
            wb = initialize_excel(wb)
            ws = wb[sheet_name]
            
            found_row_idx = None
            for r in range(2, ws.max_row + 1):
                cell_val = ws.cell(row=r, column=config['id_col_idx']).value
                if cell_val == instance.id:
                    found_row_idx = r
                    break
            
            if action == 'save':
                row_data = config['get_row_data'](instance)
                if found_row_idx:
                    for c_idx, val in enumerate(row_data, start=1):
                        ws.cell(row=found_row_idx, column=c_idx, value=val)
                    row_idx = found_row_idx
                else:
                    ws.append(row_data)
                    row_idx = ws.max_row
                
                format_row_cells(ws, row_idx, config)
            elif action == 'delete':
                if found_row_idx:
                    ws.delete_rows(found_row_idx, amount=1)
            
            adjust_column_widths(ws)
            save_workbook_with_retry(wb, FILEPATH)
    except Exception as e:
        logger.exception(f"Excel synchronization error for sheet {sheet_name} with action {action}: {str(e)}")

def export_all_to_excel():
    from home.models import Category, Product, Order, OrderItem, Cart, Wishlist
    from django.contrib.auth.models import User
    
    with excel_lock:
        wb = openpyxl.Workbook()
        wb = initialize_excel(wb)
        
        # Categories
        ws_cat = wb['Categories']
        for obj in Category.objects.all():
            ws_cat.append(SHEETS_CONFIG['Categories']['get_row_data'](obj))
            format_row_cells(ws_cat, ws_cat.max_row, SHEETS_CONFIG['Categories'])
        adjust_column_widths(ws_cat)
            
        # Products
        ws_prod = wb['Products']
        for obj in Product.objects.all():
            ws_prod.append(SHEETS_CONFIG['Products']['get_row_data'](obj))
            format_row_cells(ws_prod, ws_prod.max_row, SHEETS_CONFIG['Products'])
        adjust_column_widths(ws_prod)
            
        # Users
        ws_users = wb['Users']
        for obj in User.objects.all():
            ws_users.append(SHEETS_CONFIG['Users']['get_row_data'](obj))
            format_row_cells(ws_users, ws_users.max_row, SHEETS_CONFIG['Users'])
        adjust_column_widths(ws_users)
            
        # Orders
        ws_ord = wb['Orders']
        for obj in Order.objects.all():
            ws_ord.append(SHEETS_CONFIG['Orders']['get_row_data'](obj))
            format_row_cells(ws_ord, ws_ord.max_row, SHEETS_CONFIG['Orders'])
        adjust_column_widths(ws_ord)
            
        # Order Items
        ws_items = wb['Order_Items']
        for obj in OrderItem.objects.all():
            ws_items.append(SHEETS_CONFIG['Order_Items']['get_row_data'](obj))
            format_row_cells(ws_items, ws_items.max_row, SHEETS_CONFIG['Order_Items'])
        adjust_column_widths(ws_items)
            
        # Carts
        ws_cart = wb['Carts']
        for obj in Cart.objects.all():
            ws_cart.append(SHEETS_CONFIG['Carts']['get_row_data'](obj))
            format_row_cells(ws_cart, ws_cart.max_row, SHEETS_CONFIG['Carts'])
        adjust_column_widths(ws_cart)
            
        # Wishlists
        ws_wish = wb['Wishlists']
        for obj in Wishlist.objects.all():
            ws_wish.append(SHEETS_CONFIG['Wishlists']['get_row_data'](obj))
            format_row_cells(ws_wish, ws_wish.max_row, SHEETS_CONFIG['Wishlists'])
        adjust_column_widths(ws_wish)
            
        save_workbook_with_retry(wb, FILEPATH)
