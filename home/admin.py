from django.contrib import admin
from .models import Category, Product, Cart, Wishlist, Order, OrderItem, Coupon, ServiceablePincode
import openpyxl
from django.http import HttpResponse
from .excel_utils import SHEETS_CONFIG, format_row_cells, adjust_column_widths

def export_queryset_to_excel(modeladmin, request, queryset):
    model_name = queryset.model.__name__
    mapping = {
        'Category': 'Categories',
        'Product': 'Products',
        'User': 'Users',
        'Order': 'Orders',
        'OrderItem': 'Order_Items',
        'Cart': 'Carts',
        'Wishlist': 'Wishlists'
    }
    config_key = mapping.get(model_name)
    config = SHEETS_CONFIG.get(config_key)
    
    if not config:
        headers = [field.name for field in queryset.model._meta.fields]
        get_row_data = lambda obj: [getattr(obj, field.name) for field in queryset.model._meta.fields]
        config = {'headers': headers, 'get_row_data': get_row_data, 'id_col_idx': 1}
        
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = config_key or model_name
    
    ws.append(config['headers'])
    
    from openpyxl.styles import Font, Alignment, PatternFill
    header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='10B981', end_color='10B981', fill_type='solid')
    for col_idx in range(1, len(config['headers']) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        
    ws.freeze_panes = 'A2'
    
    for obj in queryset:
        ws.append(config['get_row_data'](obj))
        row_idx = ws.max_row
        format_row_cells(ws, row_idx, config)
        
    adjust_column_widths(ws)
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="selected_{model_name.lower()}s.xlsx"'
    wb.save(response)
    return response

export_queryset_to_excel.short_description = "Export selected items to standalone Excel file"


@admin.register(Category)
class CategoryAdmin(admin.ModelAdmin):
    list_display = ['name', 'slug', 'description']
    prepopulated_fields = {'slug': ('name',)}
    search_fields = ['name']
    actions = [export_queryset_to_excel]


class OrderItemInline(admin.TabularInline):
    model = OrderItem
    extra = 0
    readonly_fields = ['product_name', 'quantity', 'price', 'get_subtotal']

    def get_subtotal(self, obj):
        return f"₹{obj.get_subtotal()}"
    get_subtotal.short_description = 'Subtotal'


@admin.register(Product)
class ProductAdmin(admin.ModelAdmin):
    list_display = ['name', 'category', 'price', 'stock', 'rating', 'is_featured', 'created_at']
    list_filter = ['category', 'is_featured']
    search_fields = ['name', 'description']
    list_editable = ['price', 'stock', 'is_featured']
    list_per_page = 20
    actions = [export_queryset_to_excel]


@admin.register(Cart)
class CartAdmin(admin.ModelAdmin):
    list_display = ['user', 'product', 'quantity', 'get_subtotal', 'added_at']
    list_filter = ['user']
    search_fields = ['user__username', 'product__name']
    actions = [export_queryset_to_excel]

    def get_subtotal(self, obj):
        return f"₹{obj.get_subtotal()}"
    get_subtotal.short_description = 'Subtotal'


@admin.register(Wishlist)
class WishlistAdmin(admin.ModelAdmin):
    list_display = ['user', 'product', 'added_at']
    list_filter = ['user']
    search_fields = ['user__username', 'product__name']
    actions = [export_queryset_to_excel]


@admin.register(Coupon)
class CouponAdmin(admin.ModelAdmin):
    list_display = ['code', 'discount_percent', 'active', 'valid_from', 'valid_to']
    list_filter = ['active', 'valid_from', 'valid_to']
    search_fields = ['code']
    actions = [export_queryset_to_excel]


@admin.register(Order)
class OrderAdmin(admin.ModelAdmin):
    list_display = ['id', 'user', 'full_name', 'total_amount', 'discount_amount', 'status', 'created_at']
    list_filter = ['status', 'created_at']
    search_fields = ['user__username', 'full_name', 'email']
    list_editable = ['status']
    inlines = [OrderItemInline]
    readonly_fields = ['created_at', 'updated_at']
    actions = [export_queryset_to_excel]


@admin.register(OrderItem)
class OrderItemAdmin(admin.ModelAdmin):
    list_display = ['order', 'product_name', 'quantity', 'price', 'get_subtotal']
    list_filter = ['order__status']
    actions = [export_queryset_to_excel]

    def get_subtotal(self, obj):
        return f"₹{obj.get_subtotal()}"
    get_subtotal.short_description = 'Subtotal'


# Customize admin site
admin.site.site_header = "Grocery Online Admin"
admin.site.site_title = "Grocery Online"
admin.site.index_title = "Welcome to Grocery Online Admin Panel"


@admin.register(ServiceablePincode)
class ServiceablePincodeAdmin(admin.ModelAdmin):
    list_display = ['pincode', 'city', 'state']
    search_fields = ['pincode', 'city', 'state']

